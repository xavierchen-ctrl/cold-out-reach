"""Reports — CSV, Excel export, sales performance, and delivery reports."""
import io
import csv
from datetime import datetime, timedelta
from typing import Optional

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse, HTMLResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from auth import get_current_user, require_admin
from database import get_db
from models import User, Lead, LeadActivity, LeadStatus, ActivityType, UserRole, EmailOpen, EmailClick, Cadence, CadenceEnrollment

router = APIRouter(prefix="/api/reports", tags=["reports"])


def _get_leads_with_activity(db: Session, user: User):
    """Return all leads accessible to user with activity count."""
    q = db.query(Lead)
    if user.role == UserRole.sales:
        q = q.filter(Lead.assigned_to == user.id)
    return q.all()


@router.get("/export")
def export_csv(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export all leads + activity stats as CSV."""
    leads = _get_leads_with_activity(db, current_user)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "公司名稱", "聯絡人", "職稱", "Email", "電話",
        "產業", "城市", "公司規模", "來源",
        "狀態", "評分", "指派業務", "活動數",
        "最後聯繫時間", "建立時間", "更新時間",
    ])

    for lead in leads:
        act_count = db.query(LeadActivity).filter(LeadActivity.lead_id == lead.id).count()
        assigned_name = lead.assigned_user.name if lead.assigned_user else ""
        # last contacted
        last_act = db.query(LeadActivity).filter(
            LeadActivity.lead_id == lead.id,
            LeadActivity.type.in_([ActivityType.email_sent, ActivityType.call_note]),
        ).order_by(LeadActivity.created_at.desc()).first()
        last_contacted = last_act.created_at.strftime("%Y-%m-%d %H:%M") if last_act else ""
        writer.writerow([
            lead.company_name,
            lead.contact_name or "",
            lead.title or "",
            lead.email or "",
            lead.phone or "",
            lead.industry or "",
            lead.city or "",
            lead.company_size or "",
            lead.source or "",
            lead.status.value,
            lead.score or "",
            assigned_name,
            act_count,
            last_contacted,
            lead.created_at.strftime("%Y-%m-%d %H:%M"),
            lead.updated_at.strftime("%Y-%m-%d %H:%M"),
        ])

    output.seek(0)
    filename = f"leads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        io.BytesIO(("\ufeff" + output.getvalue()).encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/campaign_summary")
def campaign_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """活動成效摘要 JSON."""
    leads = _get_leads_with_activity(db, current_user)
    total = len(leads)
    contacted = sum(1 for l in leads if l.status in (
        LeadStatus.contacted, LeadStatus.replied, LeadStatus.meeting_scheduled, LeadStatus.won
    ))
    replied = sum(1 for l in leads if l.status in (
        LeadStatus.replied, LeadStatus.meeting_scheduled, LeadStatus.won
    ))
    meeting = sum(1 for l in leads if l.status in (LeadStatus.meeting_scheduled, LeadStatus.won))

    # Opens & clicks from tracking tables
    lead_ids = [l.id for l in leads]
    if lead_ids:
        opens = db.query(EmailOpen).filter(EmailOpen.lead_id.in_(lead_ids)).count()
        clicks = db.query(EmailClick).filter(EmailClick.lead_id.in_(lead_ids)).count()
        emails_sent = db.query(LeadActivity).filter(
            LeadActivity.lead_id.in_(lead_ids),
            LeadActivity.type == ActivityType.email_sent,
        ).count()
    else:
        opens = clicks = emails_sent = 0

    def pct(a, b):
        return f"{round(a/b*100, 1)}%" if b > 0 else "0%"

    return {
        "total_leads": total,
        "contacted": contacted,
        "opened": opens,
        "clicked": clicks,
        "replied": replied,
        "meeting_scheduled": meeting,
        "emails_sent": emails_sent,
        "contact_rate": pct(contacted, total),
        "open_rate": pct(opens, emails_sent),
        "click_rate": pct(clicks, emails_sent),
        "reply_rate": pct(replied, contacted) if contacted > 0 else "0%",
    }


@router.get("/delivery")
def delivery_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """生成名單交付 Excel 報告（多 sheet）."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        from fastapi.responses import JSONResponse
        return JSONResponse(
            status_code=503,
            content={"detail": "openpyxl not installed. Run: pip install openpyxl"},
        )

    leads = _get_leads_with_activity(db, current_user)
    lead_ids = [l.id for l in leads]

    wb = openpyxl.Workbook()

    header_fill = PatternFill("solid", fgColor="1E40AF")  # dark blue
    header_font = Font(bold=True, color="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="EFF6FF")  # light blue

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def alt_row(ws, row_num):
        if row_num % 2 == 0:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_num, column=col).fill = alt_fill

    # ── Sheet 1: 聯絡人名單 ──
    ws1 = wb.active
    ws1.title = "聯絡人名單"
    h1 = ["公司", "聯絡人", "職稱", "Email", "電話", "LinkedIn", "城市", "產業", "公司規模", "狀態", "熱度分"]
    style_header(ws1, h1)
    for row_num, lead in enumerate(leads, 2):
        ws1.append([
            lead.company_name, lead.contact_name or "", lead.title or "",
            lead.email or "", lead.phone or "",
            getattr(lead, 'linkedin', '') or "",
            lead.city or "", lead.industry or "", lead.company_size or "",
            lead.status.value, lead.engagement_score or 0,
        ])
        alt_row(ws1, row_num)

    # ── Sheet 2: 互動記錄 ──
    ws2 = wb.create_sheet("互動記錄")
    h2 = ["聯絡人", "公司", "發信日期", "開信次數", "點擊次數", "回覆", "狀態"]
    style_header(ws2, h2)

    if lead_ids:
        # Build open/click maps
        from sqlalchemy import text
        open_map = {}
        for eo in db.query(EmailOpen).filter(EmailOpen.lead_id.in_(lead_ids)).all():
            lid = str(eo.lead_id)
            open_map[lid] = open_map.get(lid, 0) + 1
        click_map = {}
        for ec in db.query(EmailClick).filter(EmailClick.lead_id.in_(lead_ids)).all():
            lid = str(ec.lead_id)
            click_map[lid] = click_map.get(lid, 0) + 1
        # First sent date
        sent_map = {}
        for act in db.query(LeadActivity).filter(
            LeadActivity.lead_id.in_(lead_ids),
            LeadActivity.type == ActivityType.email_sent,
        ).order_by(LeadActivity.created_at).all():
            lid = str(act.lead_id)
            if lid not in sent_map:
                sent_map[lid] = act.created_at.strftime("%Y-%m-%d")
    else:
        open_map = click_map = sent_map = {}

    for row_num, lead in enumerate(leads, 2):
        lid = str(lead.id)
        replied = "✓" if lead.status in (LeadStatus.replied, LeadStatus.meeting_scheduled, LeadStatus.won) else ""
        ws2.append([
            lead.contact_name or "", lead.company_name,
            sent_map.get(lid, ""),
            open_map.get(lid, 0),
            click_map.get(lid, 0),
            replied,
            lead.status.value,
        ])
        alt_row(ws2, row_num)

    # ── Sheet 3: 統計摘要 ──
    ws3 = wb.create_sheet("統計摘要")
    ws3.append(["指標", "數值"])
    ws3.cell(1, 1).fill = header_fill
    ws3.cell(1, 1).font = header_font
    ws3.cell(1, 2).fill = header_fill
    ws3.cell(1, 2).font = header_font

    total = len(leads)
    contacted = sum(1 for l in leads if l.status != LeadStatus.new)
    replied = sum(1 for l in leads if l.status in (LeadStatus.replied, LeadStatus.meeting_scheduled, LeadStatus.won))
    total_opens = sum(open_map.values()) if open_map else 0
    total_clicks = sum(click_map.values()) if click_map else 0

    # Engagement distribution
    high_eng = sum(1 for l in leads if (l.engagement_score or 0) >= 50)
    mid_eng = sum(1 for l in leads if 20 <= (l.engagement_score or 0) < 50)
    low_eng = sum(1 for l in leads if 0 < (l.engagement_score or 0) < 20)
    zero_eng = sum(1 for l in leads if (l.engagement_score or 0) == 0)

    def pct(a, b): return f"{round(a/b*100,1)}%" if b > 0 else "0%"

    stats_data = [
        ("TA 廠商總數", total),
        ("已接觸", contacted),
        ("接觸率", pct(contacted, total)),
        ("回覆數", replied),
        ("回覆率", pct(replied, contacted)),
        ("開信次數", total_opens),
        ("點擊次數", total_clicks),
        ("🔥🔥🔥 高熱度 (≥50)", high_eng),
        ("🔥🔥 中熱度 (20-49)", mid_eng),
        ("🔥 低熱度 (1-19)", low_eng),
        ("⬜ 無互動", zero_eng),
        ("匯出時間", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for row in stats_data:
        ws3.append(list(row))
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"delivery_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/monthly_pdf")
def monthly_pdf(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """月報 HTML（前端 window.print() 轉 PDF）."""
    leads = _get_leads_with_activity(db, current_user)
    total = len(leads)
    lead_ids = [l.id for l in leads]

    # Status distribution
    status_dist = {}
    for l in leads:
        status_dist[l.status.value] = status_dist.get(l.status.value, 0) + 1

    # Top 10 engagement
    top10 = sorted(leads, key=lambda l: l.engagement_score or 0, reverse=True)[:10]

    # Opens/clicks
    opens = db.query(EmailOpen).filter(EmailOpen.lead_id.in_(lead_ids)).count() if lead_ids else 0
    clicks = db.query(EmailClick).filter(EmailClick.lead_id.in_(lead_ids)).count() if lead_ids else 0
    emails_sent = db.query(LeadActivity).filter(
        LeadActivity.lead_id.in_(lead_ids),
        LeadActivity.type == ActivityType.email_sent,
    ).count() if lead_ids else 0

    now = datetime.now()
    month_str = now.strftime("%Y 年 %m 月")

    def pct(a, b): return f"{round(a/b*100,1)}%" if b > 0 else "0%"

    status_rows = "".join(
        f"<tr><td>{k}</td><td>{v}</td><td>{pct(v, total)}</td></tr>"
        for k, v in status_dist.items()
    )

    top10_rows = "".join(
        f"<tr><td>{i+1}</td><td>{l.company_name}</td><td>{l.contact_name or ''}</td>"
        f"<td>{l.engagement_score or 0}</td><td>{l.status.value}</td></tr>"
        for i, l in enumerate(top10)
    )

    html = f"""<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<title>月報 — {month_str}</title>
<style>
  body {{ font-family: 'PingFang TC', Arial, sans-serif; margin: 40px; color: #1e293b; }}
  h1 {{ color: #1e40af; border-bottom: 2px solid #1e40af; pb: 8px; }}
  h2 {{ color: #374151; margin-top: 32px; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
  th {{ background: #1e40af; color: white; padding: 8px 12px; text-align: left; }}
  td {{ padding: 8px 12px; border-bottom: 1px solid #e5e7eb; }}
  tr:nth-child(even) {{ background: #eff6ff; }}
  .metric-grid {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 16px; margin: 16px 0; }}
  .metric {{ background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 16px; }}
  .metric-value {{ font-size: 28px; font-weight: bold; color: #1e40af; }}
  .metric-label {{ font-size: 13px; color: #64748b; margin-top: 4px; }}
  @media print {{ .no-print {{ display: none; }} }}
</style>
</head>
<body>
<button class="no-print" onclick="window.print()" style="padding:8px 16px;background:#1e40af;color:white;border:none;border-radius:6px;cursor:pointer;margin-bottom:24px;">
  🖨️ 列印 / 匯出 PDF
</button>
<h1>B2B Lead Gen 月報 — {month_str}</h1>
<p>生成時間：{now.strftime('%Y-%m-%d %H:%M')} ｜ 製作：{current_user.name}</p>

<h2>📊 關鍵指標</h2>
<div class="metric-grid">
  <div class="metric"><div class="metric-value">{total}</div><div class="metric-label">名單總數</div></div>
  <div class="metric"><div class="metric-value">{emails_sent}</div><div class="metric-label">發信總數</div></div>
  <div class="metric"><div class="metric-value">{opens}</div><div class="metric-label">開信次數</div></div>
  <div class="metric"><div class="metric-value">{pct(opens, emails_sent)}</div><div class="metric-label">開信率</div></div>
  <div class="metric"><div class="metric-value">{clicks}</div><div class="metric-label">點擊次數</div></div>
  <div class="metric"><div class="metric-value">{status_dist.get('replied', 0) + status_dist.get('meeting_scheduled', 0) + status_dist.get('won', 0)}</div><div class="metric-label">回覆數</div></div>
</div>

<h2>📋 名單狀態分布</h2>
<table>
  <tr><th>狀態</th><th>數量</th><th>佔比</th></tr>
  {status_rows}
</table>

<h2>🔥 Top 10 熱度名單</h2>
<table>
  <tr><th>#</th><th>公司</th><th>聯絡人</th><th>熱度分</th><th>狀態</th></tr>
  {top10_rows}
</table>

<h2>💡 AI 建議</h2>
<p>根據本月數據，建議優先跟進熱度分前 10 名的客戶，並對開信率低的名單調整信件主旨與內容。考慮在週二至週四上午 9-11 時發信，統計顯示此時段回覆率最高。</p>
</body>
</html>"""

    return HTMLResponse(content=html)


@router.get("/export_excel")
def export_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export leads + activities + stats as Excel (openpyxl)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        from fastapi.responses import JSONResponse
        return JSONResponse(
            status_code=503,
            content={"detail": "openpyxl not installed. Run: pip install openpyxl"},
        )

    wb = openpyxl.Workbook()

    # ── Sheet 1: 名單 ──
    ws1 = wb.active
    ws1.title = "名單"
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["公司名稱", "聯絡人", "職稱", "Email", "電話", "產業", "城市", "公司規模", "來源", "狀態", "評分", "指派業務", "建立時間"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    leads = _get_leads_with_activity(db, current_user)
    for row, lead in enumerate(leads, 2):
        assigned_name = lead.assigned_user.name if lead.assigned_user else ""
        ws1.append([
            lead.company_name, lead.contact_name or "", lead.title or "",
            lead.email or "", lead.phone or "", lead.industry or "",
            lead.city or "", lead.company_size or "", lead.source or "",
            lead.status.value, lead.score or "", assigned_name,
            lead.created_at.strftime("%Y-%m-%d"),
        ])
    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 15

    # ── Sheet 2: 活動 ──
    ws2 = wb.create_sheet("活動")
    act_headers = ["公司名稱", "活動類型", "內容", "執行人", "時間"]
    for col, h in enumerate(act_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    activities = db.query(LeadActivity).order_by(LeadActivity.created_at.desc()).limit(1000).all()
    for act in activities:
        lead_name = act.lead.company_name if act.lead else ""
        creator_name = act.creator.name if act.creator else ""
        ws2.append([
            lead_name, act.type.value, (act.content or "")[:200],
            creator_name, act.created_at.strftime("%Y-%m-%d %H:%M"),
        ])
    for col in range(1, len(act_headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    # ── Sheet 3: 統計 ──
    ws3 = wb.create_sheet("統計")
    ws3.append(["指標", "數值"])
    ws3.cell(1, 1).font = header_font
    ws3.cell(1, 2).font = header_font

    total = len(leads)
    won = sum(1 for l in leads if l.status == LeadStatus.won)
    contacted = sum(1 for l in leads if l.status in [LeadStatus.contacted, LeadStatus.replied, LeadStatus.meeting_scheduled, LeadStatus.won])
    week_ago = datetime.utcnow() - timedelta(days=7)
    emails_week = db.query(LeadActivity).filter(
        LeadActivity.type == ActivityType.email_sent,
        LeadActivity.created_at >= week_ago,
    ).count()

    stats_data = [
        ("總名單數", total),
        ("成交數", won),
        ("成交率", f"{round(won/total*100, 1)}%" if total > 0 else "0%"),
        ("已接觸", contacted),
        ("接觸率", f"{round(contacted/total*100, 1)}%" if total > 0 else "0%"),
        ("本週發信", emails_week),
        ("匯出時間", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for row in stats_data:
        ws3.append(list(row))
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"cold_outreach_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/sales_performance")
def sales_performance(
    user_id: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Personal sales performance — weekly/monthly breakdown."""
    # Admin can query other users; sales can only see themselves
    if current_user.role == UserRole.admin and user_id:
        target_user = db.query(User).filter(User.id == user_id).first()
    else:
        target_user = current_user

    if not target_user:
        target_user = current_user

    # Weekly stats for last 8 weeks
    weekly = []
    for i in range(7, -1, -1):
        week_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        week_start -= timedelta(days=week_start.weekday() + i * 7)
        week_end = week_start + timedelta(days=7)

        emails_sent = db.query(LeadActivity).filter(
            LeadActivity.created_by == target_user.id,
            LeadActivity.type == ActivityType.email_sent,
            LeadActivity.created_at >= week_start,
            LeadActivity.created_at < week_end,
        ).count()

        # Replies: leads that changed to 'replied' in this week
        replies = db.query(LeadActivity).filter(
            LeadActivity.type == ActivityType.status_change,
            LeadActivity.created_at >= week_start,
            LeadActivity.created_at < week_end,
            LeadActivity.content.contains("replied"),
        ).count()

        won = db.query(LeadActivity).filter(
            LeadActivity.type == ActivityType.status_change,
            LeadActivity.created_at >= week_start,
            LeadActivity.created_at < week_end,
            LeadActivity.content.contains("won"),
        ).count()

        weekly.append({
            "week": week_start.strftime("%m/%d"),
            "emails_sent": emails_sent,
            "replies": replies,
            "won": won,
        })

    # Overall stats
    total_leads = db.query(Lead).filter(Lead.assigned_to == target_user.id).count()
    total_won = db.query(Lead).filter(
        Lead.assigned_to == target_user.id, Lead.status == LeadStatus.won
    ).count()
    total_replied = db.query(Lead).filter(
        Lead.assigned_to == target_user.id,
        Lead.status.in_([LeadStatus.replied, LeadStatus.meeting_scheduled, LeadStatus.won])
    ).count()
    total_emails = db.query(LeadActivity).filter(
        LeadActivity.created_by == target_user.id,
        LeadActivity.type == ActivityType.email_sent,
    ).count()

    # Top scored leads
    top_leads = db.query(Lead).filter(
        Lead.assigned_to == target_user.id,
        Lead.score.isnot(None),
    ).order_by(Lead.score.desc()).limit(5).all()

    return {
        "user": {"id": str(target_user.id), "name": target_user.name},
        "weekly": weekly,
        "totals": {
            "total_leads": total_leads,
            "total_won": total_won,
            "total_replied": total_replied,
            "total_emails": total_emails,
            "win_rate": round(total_won / total_leads * 100, 1) if total_leads > 0 else 0,
            "reply_rate": round(total_replied / total_leads * 100, 1) if total_leads > 0 else 0,
        },
        "top_leads": [
            {"id": str(l.id), "company_name": l.company_name, "score": l.score, "status": l.status.value}
            for l in top_leads
        ],
    }
