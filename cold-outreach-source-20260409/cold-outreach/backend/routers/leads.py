import csv
import io
import json
import re
from typing import Optional, List
from uuid import UUID
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
from models import User, Lead, LeadActivity, LeadStatus, ActivityType, UserRole, LeadTag, Tag, EmailOpen, EmailClick, CallLog, PendingLeadApproval
from schemas import LeadCreate, LeadUpdate, LeadStatusUpdate, LeadOut, ActivityOut
from auth import get_current_user, get_visible_user_ids
from dedup import build_conflict_index, check_conflict, add_to_index, normalize_company

router = APIRouter(prefix="/api/leads", tags=["leads"])


def _check_access(lead: Lead, user: User, db: Session):
    visible_ids = get_visible_user_ids(user, db)
    if visible_ids is not None and lead.assigned_to not in visible_ids:
        raise HTTPException(status_code=403, detail="Access denied")


# ── 名單防呆：公司名正規化 / 網域萃取，用於偵測重複或相似名單 ──────────────
_COMPANY_SUFFIXES = [
    "股份有限公司", "有限公司", "股份公司", "企業社", "工作室", "公司",
    "co.,ltd.", "co., ltd.", "co.ltd", "co ltd", "ltd.", "ltd", "inc.", "inc",
    "corporation", "corp.", "corp", "company", "co.", "group", "集團",
]


def _normalize_company(name: str) -> str:
    """去除公司型態字樣、空白與標點，用於『相似公司名』比對。"""
    n = (name or "").strip().lower()
    for suffix in _COMPANY_SUFFIXES:
        n = n.replace(suffix, "")
    # 去除台灣/分公司等地區字樣常見干擾
    n = re.sub(r"台灣|臺灣|分公司|总公司|總公司", "", n)
    n = re.sub(r"[\s\-_、,.()（）&·.]", "", n)
    return n


def _domain(url: str) -> str:
    """從官網 URL 萃取主網域（去 scheme/www/path）。"""
    if not url:
        return ""
    u = url.strip().lower()
    u = re.sub(r"^https?://", "", u)
    u = re.sub(r"^www\.", "", u)
    u = u.split("/")[0].split("?")[0].strip()
    return u


def _find_lead_conflict(body, db: Session):
    """偵測新名單是否與既有名單衝突（需人工審核）。
    回傳 (conflict_lead_id, reason) 或 (None, "")。
    觸發條件：同統一編號 / 官網網域一致 / 公司名稱相似（含同公司不同部門、不同品牌）。
    完全同公司同部門（視為單純重複）不在此攔截，維持原本可直接建立的行為。
    """
    new_norm = _normalize_company(body.company_name)
    new_tax = (getattr(body, "tax_id", "") or "").strip()
    new_domain = _domain(getattr(body, "website", "") or "")
    new_dept = (getattr(body, "department", "") or "").strip().lower()

    rows = db.query(
        Lead.id, Lead.company_name, Lead.department, Lead.tax_id, Lead.website
    ).all()

    for r in rows:
        # 1) 同一統一編號
        if new_tax and (r.tax_id or "").strip() and (r.tax_id or "").strip() == new_tax:
            return r.id, f"統一編號相同（{new_tax}）"
        # 2) 官網網域一致
        if new_domain and _domain(r.website or "") == new_domain:
            return r.id, f"官網網域相同（{new_domain}）"
        # 3) 公司名稱相似（同公司不同部門 / 不同品牌）
        if new_norm and _normalize_company(r.company_name) == new_norm:
            r_dept = (r.department or "").strip().lower()
            if r_dept == new_dept:
                continue  # 同公司同部門＝單純重複，維持原行為不攔
            return r.id, f"公司名稱相似（{r.company_name}），同公司不同部門 / 品牌"
    return None, ""


@router.get("", response_model=List[LeadOut])
def list_leads(
    status: Optional[str] = Query(None),
    assigned_to: Optional[UUID] = Query(None),
    search: Optional[str] = Query(None),
    tags: Optional[str] = Query(None),  # comma-separated tag names
    sort: Optional[str] = Query(None),  # "contact_first" = 有電話+email優先
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # 即時釋放超過 24h 未進展的「開發中」名單（→ 認領中）
    from routers.scraper import _release_expired_developments
    _release_expired_developments(db)

    q = db.query(Lead)
    # 資料權限：admin/主管=全部；小組長=自己組；業務=自己（+ 可認領的 claiming 名單池）
    visible_ids = get_visible_user_ids(current_user, db)
    if visible_ids is not None:
        # 主管=全部；小組長=組內；業務=自己（嚴格只看負責的名單）
        q = q.filter(Lead.assigned_to.in_(visible_ids))
    if status:
        q = q.filter(Lead.status == status)
    # 依業務（接洽人）篩選：admin/主管/小組長可用
    if assigned_to and current_user.role in (UserRole.admin, UserRole.manager, UserRole.team_lead):
        q = q.filter(Lead.assigned_to == assigned_to)
    if search:
        like = f"%{search}%"
        q = q.filter(
            Lead.company_name.ilike(like) |
            Lead.contact_name.ilike(like) |
            Lead.email.ilike(like)
        )
    if tags:
        tag_names = [t.strip() for t in tags.split(",") if t.strip()]
        if tag_names:
            # Filter leads that have ALL the specified tags
            for tag_name in tag_names:
                tag = db.query(Tag).filter(Tag.name == tag_name).first()
                if tag:
                    q = q.filter(
                        Lead.id.in_(
                            db.query(LeadTag.lead_id).filter(LeadTag.tag_id == tag.id)
                        )
                    )
    from sqlalchemy import case, and_
    if sort == "contact_first":
        # 有電話+email > 只有email > 只有電話 > 都沒有
        priority = case(
            (and_(Lead.phone.isnot(None), Lead.email.isnot(None)), 0),
            (and_(Lead.phone.is_(None), Lead.email.isnot(None)), 1),
            (and_(Lead.phone.isnot(None), Lead.email.is_(None)), 2),
            else_=3
        )
        q = q.order_by(priority, Lead.created_at.desc())
    else:
        q = q.order_by(Lead.created_at.desc())
    return q.offset(skip).limit(limit).all()


@router.post("")
def create_lead(
    body: LeadCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role == UserRole.sales:
        body.assigned_to = current_user.id

    # ── 名單防呆：相似公司名 / 同統編 / 同官網網域 / 同公司不同部門或品牌 → 送審 ──
    conflict_id, conflict_reason = _find_lead_conflict(body, db)
    if conflict_id:
        approval = PendingLeadApproval(
            submitted_by=current_user.id,
            lead_data=body.model_dump(mode="json"),
            conflict_company=body.company_name,
            conflict_lead_id=conflict_id,
        )
        db.add(approval)
        db.commit()
        db.refresh(approval)
        return JSONResponse(
            status_code=202,
            content={
                "pending_approval": True,
                "approval_id": str(approval.id),
                "message": (
                    f"偵測到可能重複的名單（原因：{conflict_reason}）。"
                    f"「{body.company_name}」需送請小組長及主管審核，核准後才會正式建立。"
                ),
            },
        )
    # ── 無衝突，直接建立 ──────────────────────────────────────────
    lead = Lead(**body.model_dump())
    db.add(lead)
    db.commit()
    db.refresh(lead)
    return lead


@router.get("/{lead_id}", response_model=LeadOut)
def get_lead(lead_id: UUID, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _check_access(lead, current_user, db)
    return lead


@router.patch("/{lead_id}", response_model=LeadOut)
def update_lead(
    lead_id: UUID,
    body: LeadUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _check_access(lead, current_user, db)

    old_status = lead.status
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(lead, field, value)

    if body.status and body.status != old_status:
        activity = LeadActivity(
            lead_id=lead.id,
            type=ActivityType.status_change,
            content=f"{old_status.value} → {body.status.value}",
            created_by=current_user.id,
        )
        db.add(activity)

    db.commit()
    db.refresh(lead)
    return lead


@router.delete("/{lead_id}")
def delete_lead(lead_id: UUID, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role == UserRole.sales:
        raise HTTPException(status_code=403, detail="Sales cannot delete leads")
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    db.delete(lead)
    db.commit()
    return {"message": "deleted"}


@router.patch("/{lead_id}/status", response_model=LeadOut)
def update_status(
    lead_id: UUID,
    body: LeadStatusUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _check_access(lead, current_user, db)

    old_status = lead.status
    lead.status = body.status
    activity = LeadActivity(
        lead_id=lead.id,
        type=ActivityType.status_change,
        content=f"{old_status.value} → {body.status.value}",
        created_by=current_user.id,
    )
    db.add(activity)
    db.commit()
    db.refresh(lead)
    return lead


@router.post("/recalc_engagement/all")
def recalc_engagement_all(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Recalculate engagement score for all leads."""
    leads = db.query(Lead).all()
    updated = 0
    for lead in leads:
        _recalc_one(lead, db)
        updated += 1
    db.commit()
    return {"updated": updated}


@router.post("/{lead_id}/recalc_engagement")
def recalc_engagement(
    lead_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Recalculate engagement score for a single lead."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _recalc_one(lead, db)
    db.commit()
    db.refresh(lead)
    return {"id": str(lead.id), "engagement_score": lead.engagement_score}


def _recalc_one(lead: Lead, db: Session):
    """Recalculate engagement score for a single lead object."""
    score = 0
    # Opens +10 each
    opens = db.query(EmailOpen).filter(EmailOpen.lead_id == lead.id).count()
    score += opens * 10
    # Clicks +20 each
    clicks = db.query(EmailClick).filter(EmailClick.lead_id == lead.id).count()
    score += clicks * 20
    # Replies +50
    if lead.status in ("replied", "meeting_scheduled", "won"):
        score += 50
    # Calls +30 for answered, +5 others
    calls = db.query(CallLog).filter(CallLog.lead_id == lead.id).all()
    for c in calls:
        score += 30 if c.outcome == "answered" else 5
    # Status progression +15
    activities = db.query(LeadActivity).filter(
        LeadActivity.lead_id == lead.id,
        LeadActivity.type == ActivityType.status_change,
    ).count()
    score += activities * 15
    lead.engagement_score = score


@router.get("/template")
def download_template(current_user: User = Depends(get_current_user)):
    """下載 Excel 匯入範本（含中文欄位名稱與範例資料）"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "名單範本"

    headers = ["公司名稱", "部門", "聯絡人", "職稱", "email", "電話", "產業", "城市", "公司規模", "來源"]
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 標記必填欄位 (公司名稱)
    ws.cell(row=2, column=1, value="範例公司股份有限公司")
    ws.cell(row=2, column=2, value="行銷部")
    ws.cell(row=2, column=3, value="王小明")
    ws.cell(row=2, column=4, value="行銷總監")
    ws.cell(row=2, column=5, value="example@company.com")
    ws.cell(row=2, column=6, value="02-1234-5678")
    ws.cell(row=2, column=7, value="數位行銷")
    ws.cell(row=2, column=8, value="台北")
    ws.cell(row=2, column=9, value="50-100人")
    ws.cell(row=2, column=10, value="csv_import")

    # 欄位寬度
    col_widths = [25, 12, 12, 14, 28, 16, 14, 10, 12, 14]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # 備註行
    note_cell = ws.cell(row=3, column=1, value="※ 公司名稱為必填，其餘欄位選填")
    note_cell.font = Font(color="DC2626", italic=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="lead_template.xlsx"'},
    )


@router.post("/import")
async def import_csv(
    file: UploadFile = File(...),
    check_ragic: bool = Query(False, description="是否在匯入前對 Ragic 既有客戶/陌開表去重"),
    confirmed: bool = Query(False, description="已確認衝突處理"),
    conflict_actions: Optional[str] = Query(None, description="JSON: {公司名: approve|skip}"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    content = await file.read()
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)

    # 先解析每列
    parsed: list = []
    errors: list = []
    for i, row in enumerate(rows):
        company = row.get("company_name") or row.get("公司名稱") or row.get("company")
        if not company:
            errors.append(f"Row {i+2}: missing company_name")
            continue
        parsed.append((i, row, company.strip()))

    # 若啟用 Ragic 去重，先批次查
    skipped_ragic: list = []
    ragic_lookup: dict = {}
    if check_ragic and parsed:
        try:
            from routers.ragic import _post as _ragic_post
            import asyncio as _asyncio

            async def _check(name: str):
                try:
                    e_res, n_res = await _asyncio.gather(
                        _ragic_post("get_existing_client_table", {"company_name": name}),
                        _ragic_post("get_new_client_table", {"company_name": name}),
                        return_exceptions=True,
                    )
                    e_rows = [] if isinstance(e_res, Exception) else e_res.get("data", [])
                    n_rows = [] if isinstance(n_res, Exception) else n_res.get("data", [])
                    return name, len(e_rows) > 0, len(n_rows) > 0
                except Exception:
                    return name, False, False

            sem = _asyncio.Semaphore(8)

            async def _wrapped(name):
                async with sem:
                    return await _check(name)

            unique = list({c for _, _, c in parsed})
            checks = await _asyncio.gather(*[_wrapped(n) for n in unique])
            for name, in_e, in_n in checks:
                if in_e or in_n:
                    ragic_lookup[name] = ("existing" if in_e else "new")
        except Exception as e:
            errors.append(f"Ragic 去重失敗（已忽略繼續匯入）: {e}")

    def _row_fields(row, company):
        return dict(
            company_name=company,
            department=(row.get("department") or row.get("部門") or "").strip() or None,
            contact_name=(row.get("contact_name") or row.get("聯絡人") or "").strip() or None,
            title=(row.get("title") or row.get("職稱") or "").strip() or None,
            email=(row.get("email") or row.get("Email") or "").strip() or None,
            phone=(row.get("phone") or row.get("電話") or "").strip() or None,
            industry=(row.get("industry") or row.get("產業") or "").strip() or None,
            city=(row.get("city") or row.get("城市") or "").strip() or None,
            company_size=(row.get("company_size") or row.get("公司規模") or "").strip() or None,
            source=(row.get("source") or row.get("來源") or "csv_import").strip(),
        )

    # ── 防呆：與系統現有名單比對（相似名 / 同統編 / 同網域）──────────────
    index = build_conflict_index(db)
    actions = {}
    if conflict_actions:
        try:
            actions = json.loads(conflict_actions)
        except Exception:
            actions = {}

    to_create = []      # 全新
    conflicts = []      # 需使用者決定
    skipped_dup = 0
    batch_seen = set()

    for i, row, company in parsed:
        if company in ragic_lookup:
            skipped_ragic.append({"company_name": company, "in_table": ragic_lookup[company]})
            continue
        fields = _row_fields(row, company)
        norm = normalize_company(company)
        if norm and norm in batch_seen:
            skipped_dup += 1
            continue
        # CSV 無官網欄位，以公司名 / 部門比對（同統編也無欄位）
        status, matched, reason = check_conflict(
            index, company, website=None, department=fields.get("department")
        )
        if status == "new":
            to_create.append(fields)
            if norm:
                batch_seen.add(norm)
            add_to_index(index, company, department=fields.get("department"))
        elif status == "duplicate":
            skipped_dup += 1
            if norm:
                batch_seen.add(norm)
        else:
            conflicts.append({
                "company_name": company,
                "reason": reason,
                "matched_id": matched["id"],
                "matched_company": matched["company_name"],
                "matched_department": matched.get("department"),
                "_fields": fields,
            })
            if norm:
                batch_seen.add(norm)

    # 有衝突且尚未確認 → 回傳給前端逐筆決定（不寫入）
    if conflicts and not confirmed:
        return {
            "needs_review": True,
            "conflicts": [{k: v for k, v in c.items() if k != "_fields"} for c in conflicts],
            "new_count": len(to_create),
            "skipped_ragic": skipped_ragic,
            "errors": errors,
        }

    created = 0
    pending = 0
    for fields in to_create:
        try:
            db.add(Lead(
                **fields,
                assigned_to=current_user.id if current_user.role == UserRole.sales else None,
                status=LeadStatus.new,
            ))
            created += 1
        except Exception as e:
            errors.append(f"{fields['company_name']}: {str(e)}")

    for c in conflicts:
        if actions.get(c["company_name"]) == "approve":
            db.add(PendingLeadApproval(
                submitted_by=current_user.id,
                lead_data=c["_fields"],
                conflict_company=c["company_name"],
                conflict_lead_id=c["matched_id"],
            ))
            pending += 1
        else:
            skipped_dup += 1

    db.commit()
    return {
        "created": created,
        "pending_approval": pending,
        "skipped_duplicate": skipped_dup,
        "errors": errors,
        "skipped_ragic": skipped_ragic,
    }
